"""
발견된 모순을 분류하고 메모장/엑셀로 출력
"""

from pathlib import Path
from typing import Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """검증 결과를 엑셀 파일로 생성하는 클래스"""
    
    def __init__(self, output_dir: Path):
        """
        Args:
            output_dir: 리포트를 저장할 디렉토리 경로
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_report(self, audit_results: Dict[str, Any], source_file: str) -> Path:
        """
        검증 결과를 엑셀 파일로 생성
        
        Args:
            audit_results: agent.audit_data()에서 반환된 검증 결과
            source_file: 원본 엑셀 파일명
            
        Returns:
            생성된 리포트 파일 경로
        """
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"감사_의견서_{timestamp}.xlsx"
        report_path = self.output_dir / report_filename
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "감사 의견서"
        
        # 스타일 정의
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        normal_font = Font(size=11)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        
        # 제목 행
        ws.merge_cells('A1:D1')
        ws['A1'] = "확정급여채무평가 데이터 검증 결과"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # 메타 정보
        row = 3
        ws[f'A{row}'] = "원본 파일:"
        ws[f'B{row}'] = source_file
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "검증 일시:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "검증 요약:"
        ws[f'B{row}'] = audit_results.get("summary", "")
        ws[f'A{row}'].font = Font(bold=True)
        
        # 본문 시작
        row += 3
        ws[f'A{row}'] = "검증 결과 상세"
        ws[f'A{row}'].font = title_font
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        
        # 각 시트별 검증 결과 작성
        findings = audit_results.get("findings", [])
        
        if not findings:
            row += 1
            ws[f'A{row}'] = "검증할 데이터가 없습니다."
            ws[f'A{row}'].font = normal_font
        else:
            for finding in findings:
                sheet_name = finding.get("sheet", "알 수 없음")
                result_text = finding.get("result", "")
                is_error = finding.get("error", False)
                
                # 시트명 헤더
                ws[f'A{row}'] = f"■ {sheet_name}"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws[f'A{row}'].fill = PatternFill(
                    start_color="D9E1F2", 
                    end_color="D9E1F2", 
                    fill_type="solid"
                )
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # 검증 결과 내용
                if is_error:
                    ws[f'A{row}'] = f"⚠️ 오류: {result_text}"
                    ws[f'A{row}'].font = Font(color="FF0000", bold=True)
                else:
                    # 구어체 결과를 그대로 작성
                    ws[f'A{row}'] = result_text
                    ws[f'A{row}'].font = normal_font
                
                ws[f'A{row}'].alignment = wrap_alignment
                ws.merge_cells(f'A{row}:D{row}')
                
                # 셀 높이 자동 조정을 위한 높이 설정
                ws.row_dimensions[row].height = max(60, len(result_text) // 3 * 15)
                
                row += 2
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        
        # 파일 저장
        wb.save(report_path)
        
        return report_path
    
    def generate_text_report(self, audit_results: Dict[str, Any], source_file: str) -> Path:
        """
        검증 결과를 텍스트 파일로도 생성 (추가 옵션)
        
        Args:
            audit_results: agent.audit_data()에서 반환된 검증 결과
            source_file: 원본 엑셀 파일명
            
        Returns:
            생성된 텍스트 리포트 파일 경로
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"audit_report_{timestamp}.txt"
        report_path = self.output_dir / report_filename
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("확정급여채무평가 데이터 검증 결과\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"원본 파일: {source_file}\n")
            f.write(f"검증 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"검증 요약: {audit_results.get('summary', '')}\n\n")
            f.write("-" * 60 + "\n")
            f.write("검증 결과 상세\n")
            f.write("-" * 60 + "\n\n")
            
            findings = audit_results.get("findings", [])
            for finding in findings:
                sheet_name = finding.get("sheet", "알 수 없음")
                result_text = finding.get("result", "")
                is_error = finding.get("error", False)
                
                f.write(f"\n■ {sheet_name}\n")
                if is_error:
                    f.write(f"⚠️ 오류: {result_text}\n")
                else:
                    f.write(f"{result_text}\n")
                f.write("\n" + "-" * 60 + "\n")
        
        return report_path
