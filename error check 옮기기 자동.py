"""
확정급여 채무평가 명부검증 완전 자동화 스크립트 (최종 버전)

실제 Excel 파일 구조를 반영한 완전한 자동화 스크립트
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys
import glob
import tempfile

# 스크립트 파일이 있는 디렉토리로 이동
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# ============================================================================
# 유틸리티 함수
# ============================================================================
def xls_to_xlsx(xls_파일명):
    """.xls 파일을 .xlsx로 변환"""
    임시_파일명 = xls_파일명.replace('.xls', '_temp.xlsx')
    print(f"   .xls 파일 발견 → .xlsx로 변환 중...")
    try:
        # pandas로 .xls 파일의 모든 시트 읽기
        excel_file = pd.ExcelFile(xls_파일명, engine='xlrd')
        # openpyxl 엔진으로 새 .xlsx 파일 생성
        with pd.ExcelWriter(임시_파일명, engine='openpyxl') as writer:
            for 시트명 in excel_file.sheet_names:
                df = pd.read_excel(xls_파일명, sheet_name=시트명, engine='xlrd')
                df.to_excel(writer, sheet_name=시트명, index=False)
        print(f"   ✓ 변환 완료: {임시_파일명}")
        return 임시_파일명
    except ImportError:
        print(f"✗ .xls 파일 변환 실패: xlrd 라이브러리가 필요합니다")
        print("   다음 명령어로 설치하세요: pip install xlrd")
        sys.exit(1)
    except Exception as e:
        print(f"✗ .xls 파일 변환 실패: {e}")
        sys.exit(1)

# ============================================================================
# 설정
# ============================================================================
# 파일명 자동 검색
print("=" * 80)
print("명부검증 완전 자동화 (최종 버전)")
print("=" * 80)
print("\n[0단계] 파일 검색 중...")

작성요청_파일명 = None
에러체크_파일명 = None

# 작성요청 파일 찾기 (파일명에 "확정급여채무평가" 포함)
# 단, "_자동화결과" 파일은 제외
작성요청_후보 = []
for 확장자 in ["*.xlsx", "*.xls"]:  # .xlsx와 .xls 모두 검색
    for 파일 in glob.glob(확장자):
        if "_자동화결과" in 파일 or "~$" in 파일 or "_temp" in 파일:  # 임시 파일과 결과 파일 제외
            continue
        if "확정급여채무평가" in 파일:
            작성요청_후보.append(파일)

if 작성요청_후보:
    작성요청_원본_파일명 = 작성요청_후보[0]  # 첫 번째 매칭 파일 사용
    if len(작성요청_후보) > 1:
        print(f"⚠ 여러 작성요청 파일 발견: {작성요청_후보}")
        print(f"   첫 번째 파일 사용: {작성요청_원본_파일명}")
    else:
        print(f"✓ 작성요청 파일 발견: {작성요청_원본_파일명}")
    
    # .xls 파일인 경우 .xlsx로 변환
    if 작성요청_원본_파일명.endswith('.xls') and not 작성요청_원본_파일명.endswith('.xlsx'):
        작성요청_파일명 = xls_to_xlsx(작성요청_원본_파일명)
        작성요청_원본_임시파일 = True  # 나중에 삭제할 파일인지 표시
    else:
        작성요청_파일명 = 작성요청_원본_파일명
        작성요청_원본_임시파일 = False
else:
    print("✗ 작성요청 파일을 찾을 수 없습니다.")
    print("   파일명에 '확정급여채무평가'가 포함되어 있어야 합니다.")
    print("\n현재 폴더의 Excel 파일 목록:")
    모든_파일 = [f for f in glob.glob("*.xlsx") + glob.glob("*.xls") if "_자동화결과" not in f and "~$" not in f and "_temp" not in f]
    if 모든_파일:
        for 파일 in 모든_파일:
            print(f"  - {파일}")
    else:
        print("  (Excel 파일이 없습니다)")
    sys.exit(1)

# 에러체크 파일 찾기 (파일명에 "error check" 포함)
# 단, "_자동화결과" 파일은 제외
에러체크_후보 = []
for 확장자 in ["*.xlsx", "*.xls"]:  # .xlsx와 .xls 모두 검색
    for 파일 in glob.glob(확장자):
        if "_자동화결과" in 파일 or "~$" in 파일 or "_temp" in 파일:  # 임시 파일과 결과 파일 제외
            continue
        파일_소문자 = 파일.lower()
        if "error check" in 파일_소문자:
            에러체크_후보.append(파일)

if 에러체크_후보:
    에러체크_원본_파일명 = 에러체크_후보[0]
    if len(에러체크_후보) > 1:
        print(f"⚠ 여러 에러체크 파일 발견: {에러체크_후보}")
        print(f"   첫 번째 파일 사용: {에러체크_원본_파일명}")
    else:
        print(f"✓ 에러체크 파일 발견: {에러체크_원본_파일명}")
    
    # .xls 파일인 경우 .xlsx로 변환
    if 에러체크_원본_파일명.endswith('.xls') and not 에러체크_원본_파일명.endswith('.xlsx'):
        에러체크_파일명 = xls_to_xlsx(에러체크_원본_파일명)
        에러체크_원본_임시파일 = True  # 나중에 삭제할 파일인지 표시
    else:
        에러체크_파일명 = 에러체크_원본_파일명
        에러체크_원본_임시파일 = False
else:
    print("✗ 에러체크 파일을 찾을 수 없습니다.")
    print("   파일명에 'error check'가 포함되어 있어야 합니다.")
    print("\n현재 폴더의 Excel 파일 목록:")
    모든_파일 = [f for f in glob.glob("*.xlsx") + glob.glob("*.xls") if "_자동화결과" not in f and "~$" not in f and "_temp" not in f]
    if 모든_파일:
        for 파일 in 모든_파일:
            print(f"  - {파일}")
    else:
        print("  (Excel 파일이 없습니다)")
    sys.exit(1)

# 평가일 설정
평가시작일 = "20240101"  # error check 쉬트 B1
평가종료일 = "20240331"  # error check 쉬트 D1
정년연령 = 60  # error check 쉬트 D2

# ============================================================================
# 컬럼 매핑 정의
# ============================================================================
# 작성요청 파일 - 재직자명부 (헤더: 1행, 데이터: 2행부터)
작성요청_재직자_매핑 = {
    "사원번호": "B",      # B열
    "생년월일": "C",      # C열
    "성별": "D",          # D열
    "입사일자": "E",      # E열
    "기준급여": "F",      # F열
    "당년도퇴직금추계액": "G",  # G열
    "차년도퇴직금추계액": "H",  # H열
    "종업원구분": "I",    # I열
    "중간정산기준일": "J", # J열
    "중간정산액": "K",    # K열
    "제도구분": "L",      # L열
    "적용배수": "M",      # M열 (임원배수로 복사)
    "휴직기간등차감": "N", # N열
}

# 에러체크 파일 - 재직자명부 (헤더: 1행, 데이터: 2행부터)
에러체크_재직자_매핑 = {
    "사원번호": "A",      # A열
    "성명": "B",          # B열
    "생년월일": "C",      # C열
    "성별": "D",          # D열
    "입사일자": "E",      # E열
    "기준급여": "F",      # F열
    "당년도퇴직금추계액": "G",  # G열
    "차년도퇴직금추계액": "H",  # H열
    "직종구분": "I",      # I열
    "중간정산기준일": "J", # J열
    "중간정산액": "K",    # K열
    "임원배수": "L",      # L열 (적용배수에서 복사)
}

# 작성요청 → 에러체크 매핑
재직자_데이터_매핑 = {
    "사원번호": ("B", "A"),
    "생년월일": ("C", "C"),
    "성별": ("D", "D"),
    "입사일자": ("E", "E"),
    "기준급여": ("F", "F"),
    "당년도퇴직금추계액": ("G", "G"),
    "차년도퇴직금추계액": ("H", "H"),
    "종업원구분": ("I", "I"),  # 직종구분
    "중간정산기준일": ("J", "J"),
    "중간정산액": ("K", "K"),
    "적용배수": ("M", "L"),  # 적용배수 → 임원배수
}

def excel_날짜변환(날짜값):
    """다양한 날짜 형식을 Excel 날짜 serial number로 변환"""
    if pd.isna(날짜값) or 날짜값 == "":
        return None
    
    # 이미 숫자인 경우
    if isinstance(날짜값, (int, float)):
        # Excel serial number 범위
        if 1 < 날짜값 < 100000:
            return 날짜값
        # YYYYMMDD 형식
        elif len(str(int(날짜값))) == 8:
            날짜_str = str(int(날짜값))
            try:
                dt = datetime(int(날짜_str[:4]), int(날짜_str[4:6]), int(날짜_str[6:8]))
                기준일 = datetime(1899, 12, 30)
                return (dt - 기준일).days
            except:
                return None
    
    # datetime 객체
    if isinstance(날짜값, (datetime, pd.Timestamp)):
        기준일 = datetime(1899, 12, 30)
        return (날짜값 - 기준일).days
    
    # 문자열
    if isinstance(날짜값, str):
        날짜값 = 날짜값.strip()
        try:
            dt = pd.to_datetime(날짜값)
            기준일 = datetime(1899, 12, 30)
            return (dt - 기준일).days
        except:
            return None
    
    return None

def 생년월일_연도_수정(날짜_숫자):
    """생년월일 Excel serial number에서 연도 추출하여 이상값 수정"""
    if 날짜_숫자 is None:
        return None
    
    try:
        기준일 = datetime(1899, 12, 30)
        dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
        연도 = dt.year
        
        # 1901 → 2001
        if 연도 == 1901:
            수정된_dt = dt.replace(year=2001)
            return excel_날짜변환(수정된_dt)
        
        # 2070 → 1970
        if 연도 == 2070:
            수정된_dt = dt.replace(year=1970)
            return excel_날짜변환(수정된_dt)
        
        # 1900~1905 → +100년
        if 1900 <= 연도 <= 1905:
            수정된_dt = dt.replace(year=연도 + 100)
            return excel_날짜변환(수정된_dt)
        
        # 2100 이후 → 1900년대로 변환
        if 연도 > 2100:
            수정된_dt = dt.replace(year=1900 + (연도 % 100))
            return excel_날짜변환(수정된_dt)
        
        return 날짜_숫자
    except:
        return 날짜_숫자

# ============================================================================
# 메인 프로세스
# ============================================================================
try:
    # 1. 파일 열기
    print("\n[1단계] 파일 열기...")
    # 작성요청 파일은 수식과 값을 모두 읽기 위해 두 번 열기
    작성요청_wb = load_workbook(작성요청_파일명, data_only=False)  # 수식 읽기용
    작성요청_wb_값 = load_workbook(작성요청_파일명, data_only=True)  # 계산된 값 읽기용
    에러체크_wb = load_workbook(에러체크_파일명, data_only=False)
    print("✓ 파일 열기 완료")
    
    # 2. error check 쉬트 설정
    print("\n[2단계] error check 쉬트 설정...")
    error_ws = 에러체크_wb["error check"]
    error_ws["B1"] = 평가시작일
    error_ws["D1"] = 평가종료일
    error_ws["D2"] = 정년연령
    print(f"✓ 평가시작일: {평가시작일}")
    print(f"✓ 평가종료일: {평가종료일}")
    print(f"✓ 정년연령: {정년연령}")
    
    # 2-1. 작성기준일 읽기 (작성요청 파일의 기초자료 퇴직급여 시트 I25 → 에러체크 파일 M1)
    print("\n[2-1단계] 작성기준일 읽기...")
    기초자료_시트명 = None
    for 시트명 in 작성요청_wb.sheetnames:
        if "기초자료" in 시트명 and "퇴직급여" in 시트명:
            기초자료_시트명 = 시트명
            break
    
    if 기초자료_시트명:
        print(f"   기초자료 시트 발견: {기초자료_시트명}")
        기초자료_ws = 작성요청_wb[기초자료_시트명]
        작성기준일_값 = 기초자료_ws["I25"].value
        if 작성기준일_값 is not None:
            # 날짜 형식 변환 (점 제거하여 yyyymmdd 형식으로)
            if isinstance(작성기준일_값, str):
                # 문자열인 경우: 점(.) 제거
                작성기준일_변환 = 작성기준일_값.replace(".", "").strip()
            elif isinstance(작성기준일_값, (int, float)):
                # 숫자인 경우: yyyyMMdd 형식의 숫자로 변환 (예: 20221231)
                작성기준일_변환 = str(int(작성기준일_값))
            else:
                # datetime 객체 등: yyyymmdd 형식으로 변환
                날짜_숫자 = excel_날짜변환(작성기준일_값)
                if 날짜_숫자:
                    기준일 = datetime(1899, 12, 30)
                    dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                    작성기준일_변환 = dt.strftime("%Y%m%d")
                else:
                    작성기준일_변환 = str(작성기준일_값).replace(".", "").strip()
            
            재직자명부_ws = 에러체크_wb["재직자명부"]
            재직자명부_ws["M1"] = 작성기준일_변환
            print(f"✓ 작성기준일 M1에 저장: {작성기준일_변환}")
        else:
            print("⚠ I25 셀이 비어있습니다.")
    else:
        print("⚠ 기초자료 퇴직급여 시트를 찾을 수 없습니다.")
        print(f"   사용 가능한 시트: {작성요청_wb.sheetnames}")
    
    # 3. 작성요청 파일에서 재직자명부 데이터 읽기
    print("\n[3단계] 작성요청 파일 - 재직자명부 데이터 읽기...")
    작성요청_재직자_ws = 작성요청_wb["(2-2) 재직자 명부"]
    
    # 데이터 행 개수 확인 (사원번호가 있는 행)
    작성요청_데이터_행수 = 0
    for 행 in range(2, 작성요청_재직자_ws.max_row + 1):
        사원번호 = 작성요청_재직자_ws[f"B{행}"].value
        if 사원번호 and str(사원번호).strip():
            작성요청_데이터_행수 += 1
        else:
            break
    
    print(f"✓ 재직자 수: {작성요청_데이터_행수}명")
    
    # 4. 에러체크 파일의 재직자명부 시트에 데이터 복사
    print("\n[4단계] 재직자명부 데이터 복사...")
    에러체크_재직자_ws = 에러체크_wb["재직자명부"]
    작성요청_재직자_ws_값 = 작성요청_wb_값["(2-2) 재직자 명부"]  # 계산된 값 읽기용
    
    # M1에 작성기준일 저장 (데이터 복사 전에 먼저 실행)
    if 기초자료_시트명:
        작성기준일_값 = 기초자료_ws["I25"].value
        if 작성기준일_값 is not None:
            # 날짜 형식 변환 (점 제거하여 yyyymmdd 형식으로)
            if isinstance(작성기준일_값, str):
                작성기준일_변환 = 작성기준일_값.replace(".", "").strip()
            elif isinstance(작성기준일_값, (int, float)):
                작성기준일_변환 = str(int(작성기준일_값))
            else:
                날짜_숫자 = excel_날짜변환(작성기준일_값)
                if 날짜_숫자:
                    기준일 = datetime(1899, 12, 30)
                    dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                    작성기준일_변환 = dt.strftime("%Y%m%d")
                else:
                    작성기준일_변환 = str(작성기준일_값).replace(".", "").strip()
            
            에러체크_재직자_ws["M1"] = 작성기준일_변환
            print(f"✓ 작성기준일 M1에 저장: {작성기준일_변환}")
    
    # 휴직기간 차감 열 헤더 설정 (1행)
    # X열(오차율 오른쪽) 헤더 설정
    if 에러체크_재직자_ws["X1"].value is None:
        에러체크_재직자_ws["X1"] = "휴직기간 차감"
    # AA열 헤더 설정 (지급률 수식 참조용)
    if 에러체크_재직자_ws["AA1"].value is None:
        에러체크_재직자_ws["AA1"] = "휴직기간 차감(참조용)"
    
    # 기존 데이터 초기화 (필요시)
    # 에러체크_재직자_ws.delete_rows(2, 에러체크_재직자_ws.max_row)  # 주의: 필요시에만
    
    복사_완료 = 0
    for idx in range(작성요청_데이터_행수):
        작성요청_행 = idx + 2  # 2행부터 시작
        에러체크_행 = idx + 2  # 2행부터 시작
        
        # 사원번호 확인
        사원번호 = 작성요청_재직자_ws[f"B{작성요청_행}"].value
        if not 사원번호 or not str(사원번호).strip():
            break
        
        # 데이터 복사
        # 사원번호
        에러체크_재직자_ws[f"A{에러체크_행}"] = 사원번호
        
        # 성명(B열)에도 사원번호 그대로 입력
        에러체크_재직자_ws[f"B{에러체크_행}"] = 사원번호
        
        # 생년월일 (Excel 날짜 serial number를 yyyymmdd 형식 문자열로 변환)
        # =IF(B2="","",TEXT(B2,"yyyymmdd")) 수식과 동일한 결과를 텍스트로 직접 생성
        생년월일 = 작성요청_재직자_ws[f"C{작성요청_행}"].value
        if 생년월일:
            날짜_숫자 = excel_날짜변환(생년월일)
            if 날짜_숫자:
                # 생년월일 이상값 수정
                날짜_숫자 = 생년월일_연도_수정(날짜_숫자)
                # TEXT 함수와 동일한 변환: 날짜 serial number를 yyyymmdd 형식 텍스트로 변환
                기준일 = datetime(1899, 12, 30)
                dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                yyyymmdd_텍스트 = dt.strftime("%Y%m%d")
                에러체크_재직자_ws[f"C{에러체크_행}"] = yyyymmdd_텍스트
        
        # 성별
        성별 = 작성요청_재직자_ws[f"D{작성요청_행}"].value
        if 성별 is not None:
            에러체크_재직자_ws[f"D{에러체크_행}"] = 성별
        
        # 입사일자 (Excel 날짜 serial number를 yyyymmdd 형식 문자열로 변환)
        # =IF(B2="","",TEXT(B2,"yyyymmdd")) 수식과 동일한 결과를 텍스트로 직접 생성
        입사일자 = 작성요청_재직자_ws[f"E{작성요청_행}"].value
        if 입사일자:
            날짜_숫자 = excel_날짜변환(입사일자)
            if 날짜_숫자:
                # TEXT 함수와 동일한 변환: 날짜 serial number를 yyyymmdd 형식 텍스트로 변환
                기준일 = datetime(1899, 12, 30)
                dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                yyyymmdd_텍스트 = dt.strftime("%Y%m%d")
                에러체크_재직자_ws[f"E{에러체크_행}"] = yyyymmdd_텍스트
        
        # 기준급여
        기준급여 = 작성요청_재직자_ws[f"F{작성요청_행}"].value
        if 기준급여 is not None:
            에러체크_재직자_ws[f"F{에러체크_행}"] = 기준급여
        
        # 당년도 퇴직금추계액 (계산된 값 사용)
        당년도_값 = 작성요청_재직자_ws_값[f"G{작성요청_행}"].value
        당년도_수식 = 작성요청_재직자_ws[f"G{작성요청_행}"].value
        if 당년도_값 is not None:
            에러체크_재직자_ws[f"G{에러체크_행}"] = 당년도_값
        elif isinstance(당년도_수식, str) and 당년도_수식.startswith("="):
            에러체크_재직자_ws[f"G{에러체크_행}"] = 당년도_수식
        elif 당년도_수식 is not None:
            에러체크_재직자_ws[f"G{에러체크_행}"] = 당년도_수식
        
        # 차년도 퇴직금추계액
        # 계산된 값 사용 (data_only=True로 읽은 값 - Excel에서 계산된 결과값)
        차년도_값 = 작성요청_재직자_ws_값[f"H{작성요청_행}"].value
        # 수식 확인 (data_only=False로 읽은 값)
        차년도_수식 = 작성요청_재직자_ws[f"H{작성요청_행}"].value
        
        # 차년도 값 복사 (계산된 값 우선, 없으면 수식, 없으면 None이 아닌 값)
        if 차년도_값 is not None:
            # 계산된 값이 있으면 계산된 값 사용 (0도 포함)
            에러체크_재직자_ws[f"H{에러체크_행}"] = 차년도_값
        elif 차년도_수식 is not None:
            # 계산된 값이 없고 수식이나 값이 있으면 그대로 복사
            에러체크_재직자_ws[f"H{에러체크_행}"] = 차년도_수식
        
        # 직종구분 (종업원 구분)
        직종구분 = 작성요청_재직자_ws[f"I{작성요청_행}"].value
        if 직종구분 is not None:
            에러체크_재직자_ws[f"I{에러체크_행}"] = 직종구분
        
        # 중간정산기준일 (Excel 날짜 serial number를 yyyymmdd 형식 문자열로 변환)
        중간정산기준일 = 작성요청_재직자_ws[f"J{작성요청_행}"].value
        if 중간정산기준일:
            날짜_숫자 = excel_날짜변환(중간정산기준일)
            if 날짜_숫자:
                # TEXT 함수와 동일한 변환: 날짜 serial number를 yyyymmdd 형식 텍스트로 변환
                기준일 = datetime(1899, 12, 30)
                dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                yyyymmdd_텍스트 = dt.strftime("%Y%m%d")
                에러체크_재직자_ws[f"J{에러체크_행}"] = yyyymmdd_텍스트
        
        # 중간정산액
        중간정산액 = 작성요청_재직자_ws[f"K{작성요청_행}"].value
        if 중간정산액 is not None:
            에러체크_재직자_ws[f"K{에러체크_행}"] = 중간정산액
        
        # 임원배수 (적용배수에서 복사)
        적용배수 = 작성요청_재직자_ws[f"M{작성요청_행}"].value
        if 적용배수 is not None:
            에러체크_재직자_ws[f"L{에러체크_행}"] = 적용배수
        
        # 휴직기간 차감 (작성요청 파일의 N열 → 에러체크 파일)
        # 차년도 차이(W열) 오른쪽인 X열에 휴직기간 연환산 값 저장
        # 휴직기간을 365.25로 나눈 값(연환산) 계산
        휴직기간_일수 = 작성요청_재직자_ws[f"N{작성요청_행}"].value
        if 휴직기간_일수 is not None:
            # 휴직기간이 있으면 365.25로 나눈 연환산 값 계산
            if 휴직기간_일수 != 0:
                휴직기간_연환산 = 휴직기간_일수 / 365.25
            else:
                휴직기간_연환산 = 0
        else:
            휴직기간_연환산 = 0
        
        # X열(차년도 차이 오른쪽)에 휴직기간 연환산 값 저장
        에러체크_재직자_ws[f"X{에러체크_행}"] = 휴직기간_연환산
        # AA열에도 저장 (지급률 수식에서 참조용)
        에러체크_재직자_ws[f"AA{에러체크_행}"] = 휴직기간_연환산
        
        # 지급률(S열) 수식에 휴직기간 차감 반영
        # 기존 지급률 수식을 확인하고 휴직기간(AA열)을 빼도록 수정
        기존_지급률_수식 = 에러체크_재직자_ws[f"S{에러체크_행}"].value
        if isinstance(기존_지급률_수식, str) and 기존_지급률_수식.startswith("="):
            # 기존 수식 끝에 휴직기간 차감 추가 (이미 차감이 있으면 추가하지 않음)
            if "AA" + str(에러체크_행) not in 기존_지급률_수식:
                # 수식 끝에 -AA{행} 추가
                새_지급률_수식 = 기존_지급률_수식.rstrip() + f"-AA{에러체크_행}"
                에러체크_재직자_ws[f"S{에러체크_행}"] = 새_지급률_수식
        else:
            # 수식이 없으면 기본 지급률에서 휴직기간 차감
            # 지급률은 보통 다른 열의 값을 참조하므로, 기존 수식을 확인 후 수정
            pass  # 수식이 없으면 기존 로직 유지
        
        복사_완료 += 1
    
    print(f"✓ {복사_완료}명의 데이터 복사 완료")
    print("✓ 성명(B열)에 사원번호 자동 입력 완료")
    
    # M1에 작성기준일 재확인 및 저장 (데이터 복사 후 최종 확인)
    if 기초자료_시트명:
        작성기준일_값 = 기초자료_ws["I25"].value
        if 작성기준일_값 is not None:
            # 날짜 형식 변환 (점 제거하여 yyyymmdd 형식으로)
            if isinstance(작성기준일_값, str):
                작성기준일_변환 = 작성기준일_값.replace(".", "").strip()
            elif isinstance(작성기준일_값, (int, float)):
                작성기준일_변환 = str(int(작성기준일_값))
            else:
                날짜_숫자 = excel_날짜변환(작성기준일_값)
                if 날짜_숫자:
                    기준일 = datetime(1899, 12, 30)
                    dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                    작성기준일_변환 = dt.strftime("%Y%m%d")
                else:
                    작성기준일_변환 = str(작성기준일_값).replace(".", "").strip()
            
            에러체크_재직자_ws["M1"] = 작성기준일_변환
            print(f"✓ 작성기준일 M1 최종 저장: {작성기준일_변환}")
    
    # 5. 퇴직자명부 데이터 읽기
    print("\n[5단계] 작성요청 파일 - 퇴직자명부 데이터 읽기...")
    작성요청_퇴직자_ws = 작성요청_wb["(2-3) 퇴직자 및 DC전환자 명부"]
    
    퇴직자_데이터_행수 = 0
    for 행 in range(2, 작성요청_퇴직자_ws.max_row + 1):
        사원번호 = 작성요청_퇴직자_ws[f"B{행}"].value
        if 사원번호 and str(사원번호).strip():
            퇴직자_데이터_행수 += 1
        else:
            break
    
    print(f"✓ 퇴직자 수: {퇴직자_데이터_행수}명")
    print("⚠ 퇴직자명부 복사 로직은 재직자명부와 유사하게 구현 가능합니다")
    
    # 6. 작성요청 파일 상세 검증 수행
    print("\n[6단계] 작성요청 파일 상세 검증 수행...")
    검증_결과 = []
    검증_요약 = {
        "False_수치": [],
        "정년초과자_차년도누락": [],
        "중간정산액누락": [],
        "기준급여당년도차이": [],
        "차년도누락_임원계약직": [],
        "퇴직자명부누락": [],
        "정년초과자_목록": []
    }
    
    # pandas로 데이터 읽어서 검증
    재직자_df = pd.read_excel(작성요청_파일명, sheet_name="(2-2) 재직자 명부", header=0)
    
    # 사원번호가 있는 행만 필터링
    사원번호_컬럼 = None
    for col in 재직자_df.columns:
        if "사원번호" in str(col):
            사원번호_컬럼 = col
            break
    
    if 사원번호_컬럼:
        재직자_df = 재직자_df[재직자_df[사원번호_컬럼].notna()]
    
    # 컬럼명 찾기
    생년월일_컬럼 = None
    기준급여_컬럼 = None
    당년도_컬럼 = None
    차년도_컬럼 = None
    종업원구분_컬럼 = None
    중간정산기준일_컬럼 = None
    중간정산액_컬럼 = None
    입사일자_컬럼 = None
    
    for col in 재직자_df.columns:
        col_str = str(col)
        if "생년월일" in col_str:
            생년월일_컬럼 = col
        if "기준급여" in col_str and "차" not in col_str:
            기준급여_컬럼 = col
        if "당년도" in col_str and "퇴직금" in col_str:
            당년도_컬럼 = col
        if "차년도" in col_str and "퇴직금" in col_str:
            차년도_컬럼 = col
        if ("종업원" in col_str or "직원" in col_str) and "구분" in col_str:
            종업원구분_컬럼 = col
        if "중간정산기준일" in col_str or ("중간정산" in col_str and "일" in col_str):
            중간정산기준일_컬럼 = col
        if "중간정산액" in col_str:
            중간정산액_컬럼 = col
        if "입사일자" in col_str:
            입사일자_컬럼 = col
    
    # 1. False 수치 확인
    print("   1) False 수치 확인...")
    for 시트명 in 작성요청_wb.sheetnames:
        ws = 작성요청_wb[시트명]
        for 행 in range(1, min(ws.max_row + 1, 1000)):
            for 열 in range(1, min(ws.max_column + 1, 50)):
                셀값 = ws.cell(행, 열).value
                if 셀값 is False or (isinstance(셀값, str) and 셀값.lower() == "false"):
                    검증_결과.append({
                        "유형": "False 수치 발견",
                        "시트": 시트명,
                        "셀": f"{get_column_letter(열)}{행}",
                        "사원번호": "-",
                        "내용": f"False 값이 발견되었습니다"
                    })
                    검증_요약["False_수치"].append(f"{시트명}!{get_column_letter(열)}{행}")
    
    # 2. 생년월일 이상값 체크
    print("   2) 생년월일 이상값 확인...")
    if 생년월일_컬럼 and 사원번호_컬럼:
        for idx, row in 재직자_df.iterrows():
            생년월일 = row.get(생년월일_컬럼)
            사원번호 = row.get(사원번호_컬럼)
            
            if not pd.isna(생년월일) and not pd.isna(사원번호):
                try:
                    if isinstance(생년월일, (datetime, pd.Timestamp)):
                        연도 = 생년월일.year
                        if 연도 < 1900 or 연도 > 2100:
                            검증_결과.append({
                                "유형": "생년월일 이상",
                                "시트": "재직자명부",
                                "사원번호": 사원번호,
                                "내용": f"생년월일 연도 이상: {연도}년"
                            })
                except:
                    pass
    
    # 3. 정년초과자 확인 및 차년도 퇴직금추계액 체크
    print("   3) 정년초과자 차년도 퇴직금추계액 확인...")
    평가시작일_dt = datetime.strptime(평가시작일, "%Y%m%d")
    평가종료일_dt = datetime.strptime(평가종료일, "%Y%m%d")
    평가기준일 = 평가종료일_dt  # 평가 종료일 기준으로 나이 계산
    
    if 생년월일_컬럼 and 차년도_컬럼 and 사원번호_컬럼:
        for idx, row in 재직자_df.iterrows():
            생년월일 = row.get(생년월일_컬럼)
            사원번호 = row.get(사원번호_컬럼)
            차년도 = row.get(차년도_컬럼)
            
            if not pd.isna(생년월일) and not pd.isna(사원번호):
                try:
                    if isinstance(생년월일, (datetime, pd.Timestamp)):
                        생년월일_dt = 생년월일
                    else:
                        날짜_숫자 = excel_날짜변환(생년월일)
                        if 날짜_숫자:
                            기준일 = datetime(1899, 12, 30)
                            생년월일_dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                        else:
                            continue
                    
                    나이 = (평가기준일 - 생년월일_dt).days // 365
                    
                    # 정년초과자
                    if 나이 > 정년연령:
                        검증_요약["정년초과자_목록"].append(f"사원번호 {사원번호} (나이: {나이}세)")
                        
                        # 정년초과자의 차년도 퇴직금추계액 누락 체크
                        if pd.isna(차년도):
                            검증_결과.append({
                                "유형": "차년도 퇴직금추계액 누락 (정년초과자)",
                                "시트": "재직자명부",
                                "사원번호": 사원번호,
                                "내용": f"정년초과자(나이: {나이}세)의 차년도 퇴직금추계액이 공란입니다"
                            })
                            검증_요약["정년초과자_차년도누락"].append(f"사원번호 {사원번호} (나이: {나이}세)")
                except Exception as e:
                    pass
    
    # 4. 중간정산액 확인
    print("   4) 중간정산액 확인...")
    if 중간정산기준일_컬럼 and 중간정산액_컬럼 and 사원번호_컬럼:
        for idx, row in 재직자_df.iterrows():
            중간정산기준일 = row.get(중간정산기준일_컬럼)
            중간정산액 = row.get(중간정산액_컬럼)
            사원번호 = row.get(사원번호_컬럼)
            
            if not pd.isna(중간정산기준일) and not pd.isna(사원번호):
                try:
                    if isinstance(중간정산기준일, (datetime, pd.Timestamp)):
                        중간정산일_dt = 중간정산기준일
                    else:
                        날짜_숫자 = excel_날짜변환(중간정산기준일)
                        if 날짜_숫자:
                            기준일 = datetime(1899, 12, 30)
                            중간정산일_dt = 기준일 + pd.Timedelta(days=int(날짜_숫자))
                        else:
                            continue
                    
                    # 평가년도 내에 중간정산한 경우 중간정산액 필수
                    if 평가시작일_dt <= 중간정산일_dt <= 평가종료일_dt:
                        if pd.isna(중간정산액) or 중간정산액 == 0:
                            검증_결과.append({
                                "유형": "중간정산액 누락",
                                "시트": "재직자명부",
                                "사원번호": 사원번호,
                                "내용": f"평가년도 내 중간정산자({중간정산일_dt.strftime('%Y%m%d')})인데 중간정산액이 없습니다"
                            })
                            검증_요약["중간정산액누락"].append(f"사원번호 {사원번호} (중간정산일: {중간정산일_dt.strftime('%Y%m%d')})")
                except:
                    pass
    
    # 5. 기준급여/당년도 차이 5% 이상 체크
    print("   5) 기준급여/당년도 차이 5% 이상 확인...")
    if 기준급여_컬럼 and 당년도_컬럼 and 사원번호_컬럼:
        for idx, row in 재직자_df.iterrows():
            기준급여 = row.get(기준급여_컬럼)
            당년도 = row.get(당년도_컬럼)
            사원번호 = row.get(사원번호_컬럼)
            
            if (not pd.isna(기준급여) and not pd.isna(당년도) and 
                not pd.isna(사원번호) and 기준급여 > 0):
                차이율 = abs(당년도 - 기준급여) / 기준급여 * 100
                if 차이율 >= 5:
                    검증_결과.append({
                        "유형": "기준급여/당년도 차이 5% 이상",
                        "시트": "재직자명부",
                        "사원번호": 사원번호,
                        "내용": f"차이율: {차이율:.2f}%"
                    })
                    검증_요약["기준급여당년도차이"].append(f"사원번호 {사원번호} (차이율: {차이율:.2f}%)")
    
    # 6. 차년도 퇴직금추계액 누락 체크 (임원, 계약직)
    print("   6) 차년도 퇴직금추계액 누락 확인 (임원, 계약직)...")
    if 차년도_컬럼 and 종업원구분_컬럼 and 사원번호_컬럼:
        for idx, row in 재직자_df.iterrows():
            종업원구분 = row.get(종업원구분_컬럼)
            차년도 = row.get(차년도_컬럼)
            사원번호 = row.get(사원번호_컬럼)
            
            if (not pd.isna(종업원구분) and not pd.isna(사원번호)):
                # 임원(3) 또는 계약직(4)
                if 종업원구분 in [3, 4] and pd.isna(차년도):
                    검증_결과.append({
                        "유형": "차년도 퇴직금추계액 누락 (임원/계약직)",
                        "시트": "재직자명부",
                        "사원번호": 사원번호,
                        "내용": f"임원/계약직({종업원구분})의 차년도 퇴직금추계액이 공란입니다"
                    })
                    검증_요약["차년도누락_임원계약직"].append(f"사원번호 {사원번호} (구분: {종업원구분})")
    
    # 7. 퇴직자명부 확인
    print("   7) 퇴직자명부 확인...")
    try:
        퇴직자_시트명 = None
        for 시트명 in 작성요청_wb.sheetnames:
            if "퇴직자" in 시트명 or "DC전환" in 시트명:
                퇴직자_시트명 = 시트명
                break
        
        if 퇴직자_시트명:
            퇴직자_df = pd.read_excel(작성요청_파일명, sheet_name=퇴직자_시트명, header=0)
            퇴직자_사원번호_컬럼 = None
            퇴직금_컬럼 = None
            퇴직일_컬럼 = None
            
            for col in 퇴직자_df.columns:
                col_str = str(col)
                if "사원번호" in col_str:
                    퇴직자_사원번호_컬럼 = col
                if "퇴직금" in col_str and "추계" not in col_str:
                    퇴직금_컬럼 = col
                if "퇴직일" in col_str:
                    퇴직일_컬럼 = col
            
            if 퇴직자_사원번호_컬럼:
                퇴직자_df = 퇴직자_df[퇴직자_df[퇴직자_사원번호_컬럼].notna()]
                
                for idx, row in 퇴직자_df.iterrows():
                    사원번호 = row.get(퇴직자_사원번호_컬럼)
                    
                    if 퇴직금_컬럼 and pd.isna(row.get(퇴직금_컬럼)):
                        검증_결과.append({
                            "유형": "퇴직자명부 퇴직금 누락",
                            "시트": 퇴직자_시트명,
                            "사원번호": 사원번호,
                            "내용": "퇴직금이 누락되었습니다"
                        })
                        검증_요약["퇴직자명부누락"].append(f"사원번호 {사원번호} (퇴직금 누락)")
                    
                    if 퇴직일_컬럼 and pd.isna(row.get(퇴직일_컬럼)):
                        검증_결과.append({
                            "유형": "퇴직자명부 퇴직일 누락",
                            "시트": 퇴직자_시트명,
                            "사원번호": 사원번호,
                            "내용": "퇴직일이 누락되었습니다"
                        })
                        검증_요약["퇴직자명부누락"].append(f"사원번호 {사원번호} (퇴직일 누락)")
    except Exception as e:
        print(f"   ⚠ 퇴직자명부 확인 중 오류: {e}")
    
    # 검증 결과 출력 및 파일 저장
    if 검증_결과:
        print(f"\n⚠ 검증 결과: {len(검증_결과)}건의 이슈 발견")
        for i, 결과 in enumerate(검증_결과[:20], 1):
            print(f"  {i}. [{결과['유형']}] {결과.get('시트', '')} 사원번호 {결과['사원번호']}: {결과['내용']}")
        if len(검증_결과) > 20:
            print(f"  ... 외 {len(검증_결과) - 20}건")
        
        # 검증 결과 및 추가 작업 필요사항 텍스트 파일 생성
        검증_보고서_파일명 = "검증결과_및_추가작업필요사항.txt"
        with open(검증_보고서_파일명, "w", encoding="utf-8") as f:
            f.write("=" * 80 + "\n")
            f.write("작성요청 파일 검증 결과 및 추가 작업 필요사항\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"검증 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"작성요청 파일: {작성요청_원본_파일명}\n")
            f.write(f"총 검증 이슈: {len(검증_결과)}건\n\n")
            
            f.write("=" * 80 + "\n")
            f.write("1. 검증 결과 요약\n")
            f.write("=" * 80 + "\n\n")
            
            if 검증_요약["False_수치"]:
                f.write(f"⚠ False 수치 발견: {len(검증_요약['False_수치'])}건\n")
                for item in 검증_요약["False_수치"][:10]:
                    f.write(f"   - {item}\n")
                if len(검증_요약["False_수치"]) > 10:
                    f.write(f"   ... 외 {len(검증_요약['False_수치']) - 10}건\n")
                f.write("\n")
            
            if 검증_요약["정년초과자_차년도누락"]:
                f.write(f"⚠ 정년초과자 차년도 퇴직금추계액 누락: {len(검증_요약['정년초과자_차년도누락'])}건\n")
                for item in 검증_요약["정년초과자_차년도누락"][:10]:
                    f.write(f"   - {item}\n")
                if len(검증_요약["정년초과자_차년도누락"]) > 10:
                    f.write(f"   ... 외 {len(검증_요약['정년초과자_차년도누락']) - 10}건\n")
                f.write("\n")
            
            if 검증_요약["중간정산액누락"]:
                f.write(f"⚠ 중간정산액 누락: {len(검증_요약['중간정산액누락'])}건\n")
                for item in 검증_요약["중간정산액누락"][:10]:
                    f.write(f"   - {item}\n")
                if len(검증_요약["중간정산액누락"]) > 10:
                    f.write(f"   ... 외 {len(검증_요약['중간정산액누락']) - 10}건\n")
                f.write("\n")
            
            if 검증_요약["기준급여당년도차이"]:
                f.write(f"⚠ 기준급여/당년도 차이 5% 이상: {len(검증_요약['기준급여당년도차이'])}건\n")
                f.write("   → 고객사에 확인 요청 필요\n")
                for item in 검증_요약["기준급여당년도차이"][:10]:
                    f.write(f"   - {item}\n")
                if len(검증_요약["기준급여당년도차이"]) > 10:
                    f.write(f"   ... 외 {len(검증_요약['기준급여당년도차이']) - 10}건\n")
                f.write("\n")
            
            if 검증_요약["차년도누락_임원계약직"]:
                f.write(f"⚠ 차년도 퇴직금추계액 누락 (임원/계약직): {len(검증_요약['차년도누락_임원계약직'])}건\n")
                for item in 검증_요약["차년도누락_임원계약직"][:10]:
                    f.write(f"   - {item}\n")
                if len(검증_요약["차년도누락_임원계약직"]) > 10:
                    f.write(f"   ... 외 {len(검증_요약['차년도누락_임원계약직']) - 10}건\n")
                f.write("\n")
            
            if 검증_요약["퇴직자명부누락"]:
                f.write(f"⚠ 퇴직자명부 누락 항목: {len(검증_요약['퇴직자명부누락'])}건\n")
                for item in 검증_요약["퇴직자명부누락"][:10]:
                    f.write(f"   - {item}\n")
                if len(검증_요약["퇴직자명부누락"]) > 10:
                    f.write(f"   ... 외 {len(검증_요약['퇴직자명부누락']) - 10}건\n")
                f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("2. 추가 작업 및 확인 필요사항\n")
            f.write("=" * 80 + "\n\n")
            
            f.write("ⓐ (1-2) 시트 D34, E34 셀의 연도 확인\n")
            f.write("   - 평가년도에 맞게 연도가 수정되었는지 확인\n")
            f.write("   - 종업원수에 따른 범위 확인\n\n")
            
            if 검증_요약["False_수치"]:
                f.write("ⓑ False 수치 수정 필요\n")
                f.write("   - 발견된 False 값들을 올바른 값으로 수정 요청\n\n")
            
            if 검증_요약["정년초과자_차년도누락"] or 검증_요약["차년도누락_임원계약직"]:
                f.write("ⓒ 차년도 퇴직금추계액 입력 필요\n")
                f.write("   - 임원, 계약직, 정년초과자의 차년도 퇴직금추계액 필수 입력\n")
                f.write("   - 공란이면 PUC_채무평가 파일에서 오류 발생\n")
                f.write("   - 정년초과자/계약직의 경우, 기준급여+당년도 퇴직금 추계액 합산 값으로 사용 가능\n\n")
            
            if 검증_요약["기준급여당년도차이"]:
                f.write("ⓓ 기준급여/당년도 차이 5% 이상 확인 요청\n")
                f.write("   - 고객사에 차이 발생 원인 확인 요청\n")
                f.write(f"   - 총 {len(검증_요약['기준급여당년도차이'])}건 발견\n\n")
            
            if 검증_요약["중간정산액누락"]:
                f.write("ⓔ 중간정산액 입력 필요\n")
                f.write("   - 평가년도에 중간정산한 사람은 중간정산액 반드시 필요\n\n")
            
            if 검증_요약["퇴직자명부누락"]:
                f.write("ⓕ 퇴직자명부 누락 항목 입력 필요\n")
                f.write("   - 퇴직금, 퇴직일, 사원번호 필수 입력 확인\n\n")
            
            if 검증_요약["정년초과자_목록"]:
                f.write("ⓖ 정년초과자 목록\n")
                f.write(f"   - 총 {len(검증_요약['정년초과자_목록'])}명 확인\n")
                f.write("   - 차년도 퇴직금추계액 입력 확인 필요\n\n")
            
            f.write("ⓗ 추가 확인 사항\n")
            f.write("   - 2개년도 평가 시 명부 누락 확인 (예: 25년도 퇴직했는데 24년도 명부에 없음)\n")
            f.write("   - 퇴직금제도 혼합형(DB+DC)인 경우 DB 비율 확인\n")
            f.write("   - 기타장기가 있을 경우 DC 대상자 누락 확인\n\n")
            
            f.write("=" * 80 + "\n")
            f.write("3. 상세 검증 결과\n")
            f.write("=" * 80 + "\n\n")
            
            # 유형별로 그룹화
            유형별_결과 = {}
            for 결과 in 검증_결과:
                유형 = 결과["유형"]
                if 유형 not in 유형별_결과:
                    유형별_결과[유형] = []
                유형별_결과[유형].append(결과)
            
            for 유형, 결과_목록 in 유형별_결과.items():
                f.write(f"\n[{유형}] ({len(결과_목록)}건)\n")
                f.write("-" * 80 + "\n")
                for 결과 in 결과_목록[:50]:  # 각 유형별 최대 50건
                    f.write(f"  사원번호: {결과['사원번호']}, 시트: {결과.get('시트', '-')}, {결과['내용']}\n")
                if len(결과_목록) > 50:
                    f.write(f"  ... 외 {len(결과_목록) - 50}건\n")
        
        print(f"✓ 검증 결과 및 추가작업 필요사항 저장: {검증_보고서_파일명}")
    else:
        print("\n✓ 작성요청 파일 검증 통과")
        # 검증 통과 시에도 간단한 보고서 생성
        검증_보고서_파일명 = "검증결과_및_추가작업필요사항.txt"
        with open(검증_보고서_파일명, "w", encoding="utf-8") as f:
            f.write("=" * 80 + "\n")
            f.write("작성요청 파일 검증 결과\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"검증 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"작성요청 파일: {작성요청_원본_파일명}\n\n")
            f.write("✓ 기본 검증 통과\n\n")
            f.write("추가 확인 사항:\n")
            f.write("- (1-2) 시트 D34, E34 셀의 연도 확인\n")
            f.write("- 2개년도 평가 시 명부 누락 확인\n")
            f.write("- 혼합형(DB+DC) DB 비율 확인\n")
            f.write("- 기타장기 DC 대상자 확인\n")
        print(f"✓ 검증 결과 저장: {검증_보고서_파일명}")
    
    # 7. 결과 저장
    print("\n[7단계] 결과 저장...")
    # 저장 파일명 생성 (원본 파일명 기반)
    if 에러체크_원본_임시파일:
        # 임시 파일이면 원본 파일명에서 생성
        저장_파일명 = 에러체크_원본_파일명.replace(".xls", "_자동화결과.xlsx")
    else:
        저장_파일명 = 에러체크_파일명.replace(".xlsx", "_자동화결과.xlsx")
    에러체크_wb.save(저장_파일명)
    print(f"✓ 에러체크 파일 저장: {저장_파일명}")
    
    # 원본 파일은 그대로 유지
    print("✓ 원본 파일은 변경하지 않았습니다")
    
    # 임시 파일 정리
    if 작성요청_원본_임시파일:
        try:
            os.remove(작성요청_파일명)
            print(f"✓ 임시 파일 삭제: {작성요청_파일명}")
        except Exception as e:
            print(f"⚠ 임시 파일 삭제 실패 (무시): {작성요청_파일명}")
    if 에러체크_원본_임시파일:
        try:
            os.remove(에러체크_파일명)
            print(f"✓ 임시 파일 삭제: {에러체크_파일명}")
        except Exception as e:
            print(f"⚠ 임시 파일 삭제 실패 (무시): {에러체크_파일명}")
    
except Exception as e:
    print(f"\n✗ 오류 발생: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n" + "=" * 80)
print("명부검증 자동화 완료")
print("=" * 80)
print("\n✅ 완료된 작업:")
print("1. error check 쉬트 설정값 업데이트")
print("2. 재직자명부 데이터 복사 (사원번호, 생년월일, 입사일자, 기준급여 등)")
print("3. 생년월일 이상값 자동 수정")
print("4. 기본 검증 수행 및 결과 저장")
print("\n⚠ 추가 작업 필요:")
print("1. Excel 파일을 열어 수식 계산 확인 (F9 키)")
print("2. error check 쉬트에서 에러 확인")
print("3. 퇴직자명부 데이터 복사 (필요시)")
